import os
import json
import docx
import PyPDF2
import pandas as pd
from datetime import datetime
from fireworks.client import Fireworks
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from sentence_transformers import SentenceTransformer
from qdrant_client import QdrantClient
from qdrant_client.http import models
import numpy as np
import re
import src.config as config

# ----------------------------
# Fireworks Setup - USING LLAMA FOR RELIABILITY
# ----------------------------
fw = Fireworks(api_key=config.FIREWORKS_API_KEY)

# Use Llama 3.3 70B - Much better for structured output than Qwen thinking model
LLM_MODEL = "accounts/fireworks/models/llama-v3p3-70b-instruct"

EMBED_MODEL = "sentence-transformers/multi-qa-MiniLM-L6-cos-v1"
sbert_model = SentenceTransformer(EMBED_MODEL)

# Qdrant client
qdrant = QdrantClient("http://localhost:6333")
COLLECTION_NAME = "cv_ranking"

# ----------------------------
# CV Validation Functions
# ----------------------------
def is_arabic_text(text):
    """Check if text contains significant Arabic characters"""
    if not text:
        return False
    arabic_chars = len(re.findall(r'[\u0600-\u06FF]', text))
    total_chars = len(re.findall(r'[a-zA-Z\u0600-\u06FF]', text))
    
    # If more than 30% of text is Arabic, consider it Arabic
    if total_chars > 0 and (arabic_chars / total_chars) > 0.3:
        return True
    return False

def has_sufficient_english_content(text):
    """Check if text has sufficient English content"""
    if not text:
        return False
    
    # Count English words (letters only)
    english_words = re.findall(r'\b[a-zA-Z]+\b', text)
    return len(english_words) >= 50  # At least 50 English words

def is_valid_cv_content(text):
    """
    Validate if CV has meaningful content.
    Returns: (is_valid: bool, reason: str)
    """
    if not text or len(text.strip()) < 100:
        return False, "CV is empty or too short (less than 100 characters)"
    
    # Check if it's mostly Arabic text
    if is_arabic_text(text):
        return False, "CV contains primarily Arabic text - English CV required"
    
    # Check if it has sufficient English content
    if not has_sufficient_english_content(text):
        return False, "CV lacks sufficient English content - may be an image or corrupted file"
    
    # Check for common CV indicators (at least one should be present)
    cv_indicators = [
        r'\b(experience|education|skills|qualifications|employment|work history)\b',
        r'\b(bachelor|master|phd|degree|university|college)\b',
        r'\b(hospital|clinic|medical|doctor|physician|surgeon)\b',
        r'\b(email|phone|address|contact)\b',
        r'\b(january|february|march|april|may|june|july|august|september|october|november|december|\d{4})\b'
    ]
    
    indicators_found = 0
    for pattern in cv_indicators:
        if re.search(pattern, text.lower()):
            indicators_found += 1
    
    if indicators_found < 2:
        return False, "CV does not contain expected sections (education, experience, contact info, etc.)"
    
    # Check for excessive special characters or gibberish
    special_char_ratio = len(re.findall(r'[^a-zA-Z0-9\s\.,;:\-\(\)\[\]\/]', text)) / max(len(text), 1)
    if special_char_ratio > 0.3:
        return False, "CV contains excessive special characters - may be corrupted or encoded incorrectly"
    
    # Check for repeated characters (common in corrupted files)
    repeated_patterns = re.findall(r'(.)\1{10,}', text)
    if repeated_patterns:
        return False, "CV contains suspicious repeated patterns - may be corrupted"
    
    return True, "Valid"

def validate_cv_file(cv_data):
    """
    Validate CV before processing.
    Returns: (is_valid: bool, reason: str, cleaned_text: str)
    """
    text = cv_data.get("text", "")
    filename = cv_data.get("filename", "unknown")
    
    # Basic validation
    if not text or text.strip() == "":
        return False, "Failed to extract text from CV - file may be image-based, corrupted, or password-protected", ""
    
    # Content validation
    is_valid, reason = is_valid_cv_content(text)
    
    if not is_valid:
        print(f"   ⚠️ Invalid CV detected: {reason}")
        return False, reason, ""
    
    return True, "Valid", text

# ----------------------------
# CV Parsing Functions
# ----------------------------
def extract_text_from_pdf(file_path):
    try:
        text = ""
        with open(file_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            
            # Check if PDF is encrypted
            if reader.is_encrypted:
                print(f"⚠️ PDF is password-protected: {file_path}")
                return ""
            
            for page in reader.pages:
                page_text = page.extract_text() or ""
                text += page_text
        
        return text.strip()
    except Exception as e:
        print(f"❌ PDF parse error {file_path}: {e}")
        return ""

def extract_text_from_docx(file_path):
    try:
        doc = docx.Document(file_path)
        text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
        return text.strip()
    except Exception as e:
        print(f"❌ DOCX parse error {file_path}: {e}")
        return ""

def extract_text_from_txt(file_path):
    """Extract text from TXT file."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read().strip()
    except UnicodeDecodeError:
        # Try different encoding if UTF-8 fails
        try:
            with open(file_path, 'r', encoding='latin-1') as f:
                return f.read().strip()
        except Exception as e:
            print(f"❌ TXT parse error {file_path}: {e}")
            return ""
    except Exception as e:
        print(f"❌ TXT parse error {file_path}: {e}")
        return ""

def extract_text(file_path):
    """Extract text from PDF, DOCX, DOC, or TXT files."""
    if not file_path or not os.path.exists(file_path):
        return ""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".pdf":
        return extract_text_from_pdf(file_path)
    elif ext in [".docx", ".doc"]:
        return extract_text_from_docx(file_path)
    elif ext == ".txt":
        return extract_text_from_txt(file_path)
    return ""

# ----------------------------
# NEW: Job Description Loading Function
# ----------------------------
def load_job_description(job_description_input):
    """
    Load job description from either:
    - String (direct text)
    - File path (TXT, PDF, DOCX)
    
    Args:
        job_description_input: Either a string with job description text or path to file
        
    Returns:
        str: The job description text
    """
    # If it's a file path
    if isinstance(job_description_input, str) and os.path.exists(job_description_input):
        print(f"📄 Loading job description from file: {job_description_input}")
        text = extract_text(job_description_input)
        if text:
            print(f"✅ Successfully loaded job description ({len(text)} characters)")
            return text
        else:
            print("⚠️ Warning: Could not extract text from file, using empty job description")
            return ""
    
    # If it's direct text
    elif isinstance(job_description_input, str):
        if len(job_description_input.strip()) > 0:
            print(f"📝 Using provided job description text ({len(job_description_input)} characters)")
            return job_description_input.strip()
        else:
            print("⚠️ Warning: Empty job description provided")
            return ""
    
    # Invalid input
    else:
        print("⚠️ Warning: Invalid job description input, using empty string")
        return ""

def load_cvs_from_dataframe(df):
    candidates = []
    if isinstance(df, list):
        df = pd.DataFrame(df)

    for idx, row in df.iterrows():
        cv_path = row.get("local_cv_path")
        cv_link = row.get("CV", "")
        name = row.get("Full Name") or row.get("full name") or row.get("fullname") or f"Candidate {idx+1}"

        text = ""
        if cv_path and os.path.exists(cv_path):
            try:
                text = extract_text(cv_path)
            except Exception:
                text = ""

        candidates.append({
            "filename": os.path.basename(cv_path) if cv_path else f"candidate_{idx+1}.pdf",
            "text": text,
            "name": name,
            "cv_link": cv_link
        })

    return candidates

# ----------------------------
# Ranking with Fireworks (Llama 3.3) - WITH VALIDATION
# ----------------------------
def rank_with_gemini(cvs, job_description, api_key=None, batch_size=1):
    """
    Rank CVs using Fireworks AI API with Llama 3.3 70B.
    Optimized for medical/healthcare recruitment.
    Now includes CV validation to filter out invalid CVs.
    """
    # Load job description if it's a file path
    jd_text = load_job_description(job_description)
    
    if not jd_text:
        print("⚠️ Warning: Empty job description - rankings may be inaccurate")
    
    results = []
    
    if not api_key:
        print("⚠️ No Fireworks API key found.")
        for c in cvs:
            results.append({
                "filename": c["filename"],
                "name": c["name"],
                "score": 0,
                "reasoning": "API key missing - cannot analyze",
                "cv_link": c.get("cv_link", "")
            })
        return results

    # Process CVs one at a time
    for i, cv in enumerate(cvs):
        try:
            print(f"🔥 Analyzing CV {i+1}/{len(cvs)}: {cv['name']}")
            
            # VALIDATE CV FIRST
            is_valid, validation_reason, cleaned_text = validate_cv_file(cv)
            
            if not is_valid:
                # Invalid CV - score 0
                results.append({
                    "filename": cv["filename"],
                    "name": cv["name"],
                    "score": 0,
                    "reasoning": f"❌ INVALID CV: {validation_reason}",
                    "cv_link": cv.get("cv_link", "")
                })
                print(f"   ❌ Score: 0/100 (Invalid CV)")
                print(f"   📋 Reason: {validation_reason}")
                continue
            
            # CV is valid - proceed with AI analysis
            print(f"   ✅ CV validation passed - proceeding with analysis")
            
            # Medical-specific prompt
            prompt = f"""You are an expert medical recruiter specializing in healthcare hiring. Analyze this doctor's CV against the job requirements.

JOB REQUIREMENTS:
{jd_text[:1500]}

CANDIDATE CV:
Name: {cv['name']}
{cv['text'][:2000]}

EVALUATION CRITERIA for Medical Professionals:
- Medical qualifications, degrees, and certifications
- Specialization match (e.g., Cardiology, Pediatrics, Surgery, etc.)
- Years of clinical experience and practice settings
- Specific procedures, treatments, or techniques mentioned
- Hospital affiliations and training programs
- Languages spoken (important for patient communication)
- Publications, research, or academic contributions
- Board certifications and licenses

SCORING SCALE:
- 90-100: Excellent match - specialization perfect, extensive relevant experience
- 75-89: Strong match - right specialization, good experience level
- 60-74: Good match - relevant medical background, some experience
- 40-59: Moderate match - general medical background, limited specialty match
- 20-39: Weak match - different specialization or insufficient experience
- 0-19: Poor match - unrelated medical field or entry-level when senior needed

Analyze the medical qualifications and specialization match carefully.

Respond with ONLY this JSON format:
{{"score": 85, "reasoning": "Brief explanation focusing on medical qualifications and specialty match"}}

JSON only:"""

            response = fw.chat.completions.create(
                model=LLM_MODEL,
                messages=[
                    {"role": "system", "content": "You are an expert medical recruiter with deep knowledge of healthcare qualifications and specializations. Respond only with valid JSON."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.2,
                max_tokens=400
            )

            raw_text = response.choices[0].message.content.strip()
            
            # Parse JSON
            parsed = None
            try:
                # Clean markdown if present
                clean_text = re.sub(r'```(?:json)?\s*', '', raw_text)
                clean_text = re.sub(r'```', '', clean_text).strip()
                
                # Try direct parse first
                parsed = json.loads(clean_text)
                
            except json.JSONDecodeError:
                # Extract JSON with regex
                json_match = re.search(r'\{[^{}]*"score"\s*:\s*\d+[^{}]*"reasoning"\s*:[^{}]*\}', raw_text, re.DOTALL)
                if json_match:
                    try:
                        parsed = json.loads(json_match.group(0))
                    except:
                        pass
            
            # Process result
            if parsed and "score" in parsed and "reasoning" in parsed:
                score = int(parsed["score"])
                reasoning = parsed["reasoning"].strip()
                
                # Validate score
                score = max(0, min(100, score))
                
                results.append({
                    "filename": cv["filename"],
                    "name": cv["name"],
                    "score": score,
                    "reasoning": reasoning,
                    "cv_link": cv.get("cv_link", "")
                })
                print(f"   ✅ Score: {score}/100")
                print(f"   📋 {reasoning[:80]}...")
                
            else:
                # If parsing completely fails
                print(f"   ⚠️ JSON parsing failed")
                print(f"   Raw response: {raw_text[:200]}")
                results.append({
                    "filename": cv["filename"],
                    "name": cv["name"],
                    "score": 0,
                    "reasoning": "Unable to analyze CV - please review manually",
                    "cv_link": cv.get("cv_link", "")
                })

        except Exception as e:
            print(f"   ❌ Error: {str(e)[:100]}")
            import traceback
            traceback.print_exc()
            results.append({
                "filename": cv["filename"],
                "name": cv["name"],
                "score": 0,
                "reasoning": f"Processing error: {str(e)[:80]}",
                "cv_link": cv.get("cv_link", "")
            })

    return results

def analyze_with_keywords(cv_text, job_description):
    """
    REMOVED - No keyword fallback for medical CVs.
    Medical qualifications require proper AI analysis, not keyword matching.
    """
    return 0

# ----------------------------
# Embedding-based ranking - WITH VALIDATION
# ----------------------------
def rank_with_embeddings(cvs, job_description, top_k=5):
    """
    Rank CVs using semantic embeddings.
    Now includes CV validation to filter out invalid CVs.
    """
    # Load job description if it's a file path
    jd_text = load_job_description(job_description)
    
    if not jd_text:
        print("⚠️ Warning: Empty job description - rankings may be inaccurate")
        jd_text = "General position"
    
    # Filter valid CVs
    valid_cvs = []
    invalid_results = []
    
    for cv in cvs:
        is_valid, validation_reason, cleaned_text = validate_cv_file(cv)
        if is_valid:
            valid_cvs.append(cv)
        else:
            invalid_results.append({
                "filename": cv["filename"],
                "name": cv["name"],
                "score": 0,
                "reasoning": f"INVALID CV: {validation_reason}",
                "cv_link": cv.get("cv_link", "")
            })
    
    if len(valid_cvs) == 0:
        print("⚠️ No valid CVs found for embedding-based ranking")
        return invalid_results
    
    if not qdrant.collection_exists(COLLECTION_NAME):
        qdrant.recreate_collection(
            collection_name=COLLECTION_NAME,
            vectors_config=models.VectorParams(size=384, distance=models.Distance.COSINE),
        )

    qdrant.delete_collection(COLLECTION_NAME)
    qdrant.recreate_collection(
        collection_name=COLLECTION_NAME,
        vectors_config=models.VectorParams(size=384, distance=models.Distance.COSINE),
    )

    vectors = []
    for idx, c in enumerate(valid_cvs):
        emb = sbert_model.encode(c["text"] or "")
        vectors.append(models.PointStruct(
            id=idx,
            vector=emb.tolist(),
            payload={"filename": c["filename"], "name": c["name"], "cv_link": c["cv_link"], "text": c["text"]}
        ))
    qdrant.upsert(collection_name=COLLECTION_NAME, points=vectors)

    job_emb = sbert_model.encode(jd_text).tolist()
    search_results = qdrant.search(
        collection_name=COLLECTION_NAME,
        query_vector=job_emb,
        limit=min(top_k, len(valid_cvs))
    )

    results = []
    for r in search_results:
        payload = r.payload
        score = round(r.score * 100, 2)
        results.append({
            "filename": payload.get("filename", "Unknown"),
            "name": payload.get("name", "Unknown"),
            "score": score,
            "reasoning": f"Semantic similarity: {score}%",
            "cv_link": payload.get("cv_link", "")
        })
    
    # Combine valid results with invalid ones
    return results + invalid_results

# ----------------------------
# Save to Excel
# ----------------------------
def save_results_to_excel(results_list, job_description=None, output_dir=".", output_path=None):
    """
    Save ranking results to Excel file.
    """
    df_out = pd.DataFrame(results_list)

    required_cols = ["name", "score", "status", "reasoning", "cv_link"]
    for col in required_cols:
        if col not in df_out.columns:
            df_out[col] = ""

    df_out["status"] = df_out["score"].apply(lambda x: "Match" if float(x) >= 60 else "Not Match")
    df_out = df_out[required_cols].sort_values(by="score", ascending=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranked Candidates"

    headers = list(df_out.columns)
    ws.append(headers)
    for _, row in df_out.iterrows():
        ws.append([row.get(h, "") for h in headers])

    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)), start=2):
        status = ws[f"C{i}"].value
        try:
            score = float(ws[f"B{i}"].value)
        except:
            score = 0

        if status == "Match":
            fill = PatternFill(start_color="C6EFCE", fill_type="solid")
        elif score >= 40:
            fill = PatternFill(start_color="FFEB9C", fill_type="solid")
        else:
            fill = PatternFill(start_color="F2DCDB", fill_type="solid")

        for cell in row:
            cell.fill = fill

        cv_link = ws[f"E{i}"].value
        if cv_link and str(cv_link).startswith("http"):
            ws[f"E{i}"].hyperlink = cv_link
            ws[f"E{i}"].style = "Hyperlink"

    ws_summary = wb.create_sheet("Summary")
    avg_score = df_out["score"].astype(float).mean() if not df_out.empty else 0
    valid_cvs = len(df_out[df_out["score"] > 0])
    invalid_cvs = len(df_out[df_out["score"] == 0])
    top_candidate = df_out.iloc[0] if not df_out.empty else None
    
    ws_summary["A1"], ws_summary["B1"] = "Total CVs", len(df_out)
    ws_summary["A2"], ws_summary["B2"] = "Valid CVs", valid_cvs
    ws_summary["A3"], ws_summary["B3"] = "Invalid CVs", invalid_cvs
    ws_summary["A4"], ws_summary["B4"] = "Average Score (Valid)", avg_score
    
    if top_candidate is not None:
        ws_summary["A6"], ws_summary["B6"] = "Top Candidate", top_candidate.get("name", "Unknown")
        ws_summary["A7"], ws_summary["B7"] = "Top Score", top_candidate.get("score", 0)
        ws_summary["A8"], ws_summary["B8"] = "Top CV Link", top_candidate.get("cv_link", "")

    if output_path:
        final_path = output_path
    else:
        # Generate filename from job description
        if job_description and os.path.exists(job_description):
            # If it's a file, use the filename without extension
            job_name = os.path.splitext(os.path.basename(job_description))[0]
        elif job_description:
            # If it's text, use first 50 chars
            job_name = job_description[:50]
        else:
            job_name = "Job"
        
        safe_job_desc = "".join(c if c.isalnum() or c in (" ", "_", "-") else "_" for c in job_name)
        output_filename = f"{safe_job_desc.strip().replace(' ', '_')}_Ranked_Candidates.xlsx"
        final_path = os.path.join(output_dir, output_filename)

    wb.save(final_path)
    print(f"✅ Results saved to {final_path}")
    print(f"📊 Summary: {valid_cvs} valid CVs, {invalid_cvs} invalid CVs")
    return df_out, final_path